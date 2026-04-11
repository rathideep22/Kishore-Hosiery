from fastapi import FastAPI, APIRouter, HTTPException, WebSocket, WebSocketDisconnect, Depends, Header, Query, Request, UploadFile, File
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from bson.objectid import ObjectId
import os
import re
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict
from urllib.parse import unquote
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import asyncio
from pdf_utils import OrderPDFGenerator

ROOT_DIR = Path(__file__).parent
env_path = ROOT_DIR / '.env'
load_dotenv(env_path, override=True)

mongo_url = os.environ.get('MONGODB_URL')
if not mongo_url:
    raise ValueError("MONGODB_URL not found in environment. Please set it in .env or as an environment variable.")
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'kishore_hosiery')]

JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key-change-in-production')
JWT_ALGORITHM = 'HS256'
MOCK_OTP = '1234'

# AWS S3 Configuration
AWS_ACCESS_KEY_ID = os.environ.get('AWS_ACCESS_KEY_ID')
AWS_SECRET_ACCESS_KEY = os.environ.get('AWS_SECRET_ACCESS_KEY')
AWS_S3_BUCKET = os.environ.get('AWS_S3_BUCKET', 'bills-kishore')
AWS_REGION = os.environ.get('AWS_REGION', 'us-east-1')

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# ─── Models ───────────────────────────────────────────────────────────────────

class SendOTPRequest(BaseModel):
    phone: str

class VerifyOTPRequest(BaseModel):
    phone: str
    otp: str

class CreateUserRequest(BaseModel):
    phone: str
    firstName: str
    lastName: str
    role: str

class OrderProductItem(BaseModel):
    productId: str
    alias: str
    category: str
    size: str
    printName: str
    quantity: int
    rate: Optional[str] = None
    requireSerialNo: bool = False
    serialNumbers: Optional[List[Optional[str]]] = None

class UpdateOrderRequest(BaseModel):
    partyName: Optional[str] = None
    location: Optional[str] = None
    godown: Optional[str] = None
    message: Optional[str] = None
    totalParcels: Optional[int] = None
    items: Optional[List[OrderProductItem]] = None

class UpdateBillRequest(BaseModel):
    billNo: str

class GodownUpdateRequest(BaseModel):
    godown: str
    readyParcels: int

class GodownPrefixRequest(BaseModel):
    prefix: str

class CreateProductRequest(BaseModel):
    category: str
    size: str
    printName: str
    alias: str

class UpdateProductRequest(BaseModel):
    category: Optional[str] = None
    size: Optional[str] = None
    printName: Optional[str] = None
    alias: Optional[str] = None

class CreateOrderRequest(BaseModel):
    partyName: str
    location: str
    message: str
    godown: str  # Which gowdown this order is for
    items: Optional[List[OrderProductItem]] = []
    totalParcels: Optional[int] = None

class ParcelFulfillmentRequest(BaseModel):
    productId: str
    parcelIndex: int
    weight: Optional[float] = None
    serialNo: Optional[str] = None

# ─── WebSocket Manager ───────────────────────────────────────────────────────

class ConnectionManager:
    def __init__(self):
        # One user may have several live sockets (phone + tablet + web),
        # so we keep a list per user_id instead of a single socket. A bare
        # dict[user_id] = socket would let the newest device silently
        # replace older ones — the original cause of the "sound only on
        # the host phone" bug.
        self.active_connections: Dict[str, List[WebSocket]] = {}

    async def connect(self, websocket: WebSocket, user_id: str):
        await websocket.accept()
        self.active_connections.setdefault(user_id, []).append(websocket)

    def disconnect(self, user_id: str, websocket: Optional[WebSocket] = None):
        sockets = self.active_connections.get(user_id)
        if not sockets:
            return
        if websocket is None:
            self.active_connections.pop(user_id, None)
            return
        try:
            sockets.remove(websocket)
        except ValueError:
            pass
        if not sockets:
            self.active_connections.pop(user_id, None)

    async def broadcast(self, message: dict):
        for uid, sockets in list(self.active_connections.items()):
            for conn in list(sockets):
                try:
                    await conn.send_json(message)
                except Exception:
                    self.disconnect(uid, conn)

    async def send_to_user(self, user_id: str, message: dict):
        for conn in list(self.active_connections.get(user_id, [])):
            try:
                await conn.send_json(message)
            except Exception:
                self.disconnect(user_id, conn)

manager = ConnectionManager()


# ─── PDF Regeneration Helpers ────────────────────────────────────────────────

# Order states in which the PDF bill is meaningful and should be kept fresh.
# Pending / Partial Ready are excluded because the data is incomplete.
PDF_ELIGIBLE_STATUSES = {"Ready", "Completed", "Bill Generated"}


def _derive_pdf_filename(order: dict) -> str:
    """Return the S3 filename (without extension) for this order's PDF.

    If the order already has a billPdfUrl, we reuse the same filename so
    the regeneration overwrites the existing S3 key — keeping one PDF per
    order instead of accumulating dated duplicates. First-time generation
    uses the current UTC date.
    """
    party_name = order.get('partyName') or 'Order'
    existing_url = order.get('billPdfUrl')
    if existing_url:
        match = re.search(r'/bills/(.+?)\.pdf', existing_url)
        if match:
            return unquote(match.group(1))
    pdf_date = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    return f"{party_name} ({pdf_date})"


async def regenerate_order_pdf(order: dict) -> Optional[str]:
    """Regenerate the PDF bill for an order and upload it to S3.

    Runs the synchronous PDF + S3 work in a threadpool to keep the
    FastAPI event loop responsive. Returns the new PDF URL, or None if
    the order is not in a PDF-eligible state or generation failed.
    """
    status = order.get('readinessStatus', 'Pending')
    if status not in PDF_ELIGIBLE_STATUSES:
        return None
    try:
        filename = _derive_pdf_filename(order)
        generator = OrderPDFGenerator()
        url = await asyncio.to_thread(
            generator.generate_order_bill, order, filename
        )
        logging.info(
            f"PDF regenerated for order {order.get('orderId')} [{status}]: {url}"
        )
        return url
    except Exception as e:
        logging.error(
            f"Failed to regenerate PDF for order {order.get('orderId')}: {e}",
            exc_info=True,
        )
        return None


async def refresh_pdf_for_order(order: dict) -> dict:
    """Regenerate the PDF for the given order and persist billPdfUrl.

    Mutates and returns the supplied order dict so callers can broadcast
    the latest state without an extra DB read. No-op for orders whose
    status isn't PDF-eligible.
    """
    pdf_url = await regenerate_order_pdf(order)
    if pdf_url:
        await db.orders.update_one(
            {"id": order["id"]},
            {"$set": {"billPdfUrl": pdf_url}},
        )
        order['billPdfUrl'] = pdf_url
    return order


# ─── Auth Helpers ─────────────────────────────────────────────────────────────

def create_token(user: dict) -> str:
    payload = {
        "userId": user['id'],
        "phone": user['phone'],
        "role": user['role'],
        "exp": datetime.now(timezone.utc) + timedelta(days=30)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_auth_user(authorization: str = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="No valid token provided")
    token = authorization.split(" ", 1)[1]
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload['userId']}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")


def require_admin(user: dict):
    if user.get('role') != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")

def require_admin_or_staff(user: dict):
    if user.get('role') not in ['admin', 'staff']:
        raise HTTPException(status_code=403, detail="Admin or staff access required")

def require_admin_or_accountant(user: dict):
    if user.get('role') not in ['admin', 'accountant']:
        raise HTTPException(status_code=403, detail="Admin or accountant access required")

# ─── Notification & Audit Helpers ─────────────────────────────────────────────

async def create_notification(user_id: str, message: str, ntype: str, order_id: str = None):
    notif = {
        "id": str(uuid.uuid4()),
        "userId": user_id,
        "message": message,
        "type": ntype,
        "orderId": order_id,
        "read": False,
        "sound": True,
        "createdAt": datetime.now(timezone.utc).isoformat()
    }
    await db.notifications.insert_one({**notif})
    await manager.send_to_user(user_id, {"type": "NOTIFICATION", "notification": notif, "sound": True})
    return notif


async def create_audit_log(user_id: str, action: str, order_id: str = None, details: str = ""):
    log_entry = {
        "id": str(uuid.uuid4()),
        "userId": user_id,
        "action": action,
        "orderId": order_id,
        "details": details,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    await db.audit_logs.insert_one({**log_entry})


# Fallback prefix when a godown isn't set or hasn't been configured yet.
DEFAULT_ORDER_PREFIX = "KH"


async def get_next_order_id(godown_name: Optional[str] = None) -> tuple[str, int]:
    """Reserve the next order number for a given godown.

    Looks up the godown's configured prefix (set via seeding or the
    update-prefix endpoint) and atomically increments a per-prefix
    counter so each warehouse has its own independent numbering. Falls
    back to DEFAULT_ORDER_PREFIX when the godown is missing or has no
    prefix — this keeps legacy orders flowing through the same path.
    """
    prefix = DEFAULT_ORDER_PREFIX
    if godown_name:
        godown = await db.gowdowns.find_one({"name": godown_name})
        if godown and godown.get("prefix"):
            prefix = godown["prefix"]
    counter_name = f"orderId:{prefix}"
    counter = await db.counters.find_one_and_update(
        {"name": counter_name},
        {"$inc": {"value": 1}},
        upsert=True,
        return_document=True,
    )
    return prefix, counter["value"]


# ─── Health Check (Public) ────────────────────────────────────────────────────

@api_router.get("/health")
async def health_check():
    return {"status": "ok", "message": "Backend is running"}


# ─── Auth Routes ──────────────────────────────────────────────────────────────

@api_router.post("/auth/send-otp")
async def send_otp(req: SendOTPRequest):
    user = await db.users.find_one({"phone": req.phone}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found. Only pre-registered users can login.")
    # Mock OTP — in production replace with Twilio
    await db.otp_store.update_one(
        {"phone": req.phone},
        {"$set": {"otp": MOCK_OTP, "createdAt": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    logger.info(f"OTP requested for pre-registered user: {req.phone}")
    return {"message": "OTP sent successfully", "mock_otp": MOCK_OTP}


@api_router.post("/auth/verify-otp")
async def verify_otp(req: VerifyOTPRequest):
    user = await db.users.find_one({"phone": req.phone}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found. Only pre-registered users can login.")
    if req.otp != MOCK_OTP:
        stored = await db.otp_store.find_one({"phone": req.phone})
        if not stored or stored.get('otp') != req.otp:
            raise HTTPException(status_code=400, detail="Invalid OTP")
    token = create_token(user)
    logger.info(f"User logged in: {req.phone} (ID: {user['id']})")
    return {"token": token, "user": user}


@api_router.get("/auth/me")
async def get_me(user: dict = Depends(get_auth_user)):
    return user


# ─── User Routes ──────────────────────────────────────────────────────────────

@api_router.get("/users")
async def get_users(user: dict = Depends(get_auth_user)):
    require_admin(user)
    users = await db.users.find({}, {"_id": 0}).to_list(100)
    return users


@api_router.post("/users")
async def create_user(req: CreateUserRequest, user: dict = Depends(get_auth_user)):
    require_admin(user)
    existing = await db.users.find_one({"phone": req.phone})
    if existing:
        raise HTTPException(status_code=400, detail="Phone number already registered")
    new_user = {
        "id": str(uuid.uuid4()),
        "phone": req.phone,
        "firstName": req.firstName,
        "lastName": req.lastName,
        "role": req.role,
        "createdAt": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one({**new_user})
    await create_audit_log(user['id'], "USER_CREATED", None, f"Created user {req.firstName} {req.lastName}")
    return new_user


@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, user: dict = Depends(get_auth_user)):
    require_admin(user)
    target = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")
    if target['role'] == 'admin':
        raise HTTPException(status_code=400, detail="Cannot delete admin users")
    await db.users.delete_one({"id": user_id})
    await create_audit_log(user['id'], "USER_DELETED", None, f"Deleted user {target['firstName']} {target['lastName']}")
    return {"message": "User deleted"}


# ─── Order Routes ─────────────────────────────────────────────────────────────

@api_router.get("/orders")
async def get_orders(
    user: dict = Depends(get_auth_user),
    status: Optional[str] = None,
    search: Optional[str] = None,
    godown: Optional[str] = None,
    include_completed: Optional[bool] = False,
    limit: int = Query(500, ge=1, le=1000),
    skip: int = Query(0, ge=0),
):
    conditions = []
    user_role = user.get('role', 'staff')

    # Exclude completed orders from normal views (unless explicitly requested)
    if not include_completed:
        conditions.append({"completed": {"$ne": True}})

    # Accountant: sees Ready/Partial Ready orders (ready to bill) or dispatched orders
    if user_role == 'accountant':
        conditions.append({"$or": [
            {"readinessStatus": {"$in": ["Ready", "Partial Ready"]}},
            {"dispatched": True}
        ]})
    elif user_role == 'staff':
        # Staff sees only non-dispatched orders
        conditions.append({"dispatched": False})
    # Admin sees all non-completed orders (no filter)

    if status == 'ready':
        conditions.append({"readinessStatus": "Ready", "dispatched": False})
    elif status == 'partial':
        conditions.append({"readinessStatus": "Partial Ready", "dispatched": False})
    elif status == 'pending':
        conditions.append({"readinessStatus": "Pending", "dispatched": False})
    elif status == 'dispatched':
        conditions.append({"dispatched": True})
    elif status == 'active':
        conditions.append({"dispatched": False})
    if search:
        # Escape regex metacharacters in the user-supplied search so a
        # party name like "A+B" doesn't explode Mongo's regex engine.
        safe_search = re.escape(search)
        conditions.append({
            "$or": [
                {"partyName": {"$regex": safe_search, "$options": "i"}},
                {"orderId": {"$regex": safe_search, "$options": "i"}}
            ]
        })
    if godown:
        conditions.append({"godown": godown})
    query = {"$and": conditions} if conditions else {}
    # Response shape stays an array (all existing callers expect that),
    # but skip/limit are now honourable so older clients keep getting the
    # newest 500 while new clients can page deeper.
    orders = await (
        db.orders.find(query, {"_id": 0})
        .sort("createdAt", -1)
        .skip(skip)
        .limit(limit)
        .to_list(length=limit)
    )
    return orders


@api_router.post("/orders")
async def create_order(req: CreateOrderRequest, user: dict = Depends(get_auth_user)):
    require_admin_or_staff(user)
    prefix, order_num = await get_next_order_id(req.godown)
    items = [item.model_dump() for item in req.items] if req.items else []
    total_parcels = sum(item.quantity for item in req.items) if req.items else req.totalParcels or 0
    order = {
        "id": str(uuid.uuid4()),
        "orderId": f"{prefix}-{order_num:04d}",
        "partyName": req.partyName,
        "location": req.location,
        "message": req.message,
        "godown": req.godown,
        "items": items,
        "totalParcels": total_parcels,
        "godownDistribution": [],
        "readinessStatus": "Pending",
        "dispatched": False,
        "dispatchedAt": None,
        "billNo": None,
        "completed": False,
        "completedAt": None,
        "createdBy": user['id'],
        "createdByName": f"{user['firstName']} {user['lastName']}",
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "updatedAt": datetime.now(timezone.utc).isoformat()
    }
    await db.orders.insert_one({**order})

    # Notify every staff / admin / accountant (except the creator) so their
    # phones chime. We skip the creator so they don't hear their own order.
    recipients = await db.users.find(
        {"role": {"$in": ["staff", "admin", "accountant"]}, "id": {"$ne": user['id']}},
        {"_id": 0},
    ).to_list(500)
    for recipient in recipients:
        await create_notification(
            recipient['id'],
            f"New Order {order['orderId']} for {order['partyName']}",
            "new_order",
            order['id']
        )
    await create_audit_log(user['id'], "ORDER_CREATED", order['id'], f"Created order {order['orderId']}")
    await manager.broadcast({"type": "ORDER_CREATED", "order": order})
    return order


@api_router.get("/orders/{order_id}")
async def get_order(order_id: str, user: dict = Depends(get_auth_user)):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    return order


@api_router.put("/orders/{order_id}")
async def update_order(order_id: str, req: UpdateOrderRequest, user: dict = Depends(get_auth_user)):
    require_admin(user)
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    update = {"updatedAt": datetime.now(timezone.utc).isoformat()}
    if req.partyName is not None:
        update["partyName"] = req.partyName
    if req.location is not None:
        update["location"] = req.location
    if req.godown is not None:
        update["godown"] = req.godown
    if req.message is not None:
        update["message"] = req.message
    if req.items is not None:
        # Merge incoming items with existing ones so that weights/serial
        # numbers already captured for unchanged products are preserved.
        # Editing an order means adding, removing, or tweaking lines — it
        # must never wipe fulfilment progress that has already happened.
        existing_by_pid = {
            (it.get('productId') or ''): it
            for it in (order.get('items') or [])
        }
        merged_items: List[dict] = []
        for incoming in req.items:
            new_item = incoming.model_dump()
            prev = existing_by_pid.get(new_item['productId'])
            new_qty = int(new_item.get('quantity') or 0)
            if prev:
                prev_fulfil = list(prev.get('fulfillment') or [])
                if len(prev_fulfil) < new_qty:
                    prev_fulfil.extend([None] * (new_qty - len(prev_fulfil)))
                else:
                    prev_fulfil = prev_fulfil[:new_qty]
                new_item['fulfillment'] = prev_fulfil

                prev_serials = list(prev.get('serialNumbers') or [])
                if len(prev_serials) < new_qty:
                    prev_serials.extend([None] * (new_qty - len(prev_serials)))
                else:
                    prev_serials = prev_serials[:new_qty]
                new_item['serialNumbers'] = prev_serials
            else:
                new_item['fulfillment'] = [None] * new_qty
                new_item['serialNumbers'] = [None] * new_qty
            merged_items.append(new_item)
        update["items"] = merged_items

        # Recalculate totalParcels and readinessStatus from the merged state
        # so status stays consistent with the new quantities.
        new_total = sum(int(it.get('quantity') or 0) for it in merged_items)
        update["totalParcels"] = new_total
        fulfilled_count = sum(
            1
            for it in merged_items
            for w in (it.get('fulfillment') or [])
            if w is not None
        )
        if new_total == 0:
            update["readinessStatus"] = "Pending"
        elif fulfilled_count >= new_total:
            update["readinessStatus"] = "Ready"
        elif fulfilled_count > 0:
            update["readinessStatus"] = "Partial Ready"
        else:
            update["readinessStatus"] = "Pending"
    elif req.totalParcels is not None:
        update["totalParcels"] = req.totalParcels
        total_ready = sum(g.get('readyParcels', 0) for g in order.get('godownDistribution', []))
        if total_ready >= req.totalParcels:
            update["readinessStatus"] = "Ready"
        elif total_ready > 0:
            update["readinessStatus"] = "Partial Ready"
        else:
            update["readinessStatus"] = "Pending"
    await db.orders.update_one({"id": order_id}, {"$set": update})
    updated = await db.orders.find_one({"id": order_id}, {"_id": 0})
    updated = await refresh_pdf_for_order(updated)
    await create_audit_log(user['id'], "ORDER_UPDATED", order_id, f"Updated order {order['orderId']}")
    await manager.broadcast({"type": "ORDER_UPDATED", "order": updated})
    return updated


@api_router.put("/orders/{order_id}/bill")
async def update_bill(order_id: str, req: UpdateBillRequest, user: dict = Depends(get_auth_user)):
    require_admin_or_accountant(user)
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    update = {
        "billNo": req.billNo,
        "readinessStatus": "Bill Generated",
        "updatedAt": datetime.now(timezone.utc).isoformat()
    }
    await db.orders.update_one({"id": order_id}, {"$set": update})
    updated = await db.orders.find_one({"id": order_id}, {"_id": 0})
    updated = await refresh_pdf_for_order(updated)

    # Notify admins when accountant adds a bill
    if user.get('role') == 'accountant':
        admins = await db.users.find({"role": "admin"}, {"_id": 0}).to_list(10)
        for admin_user in admins:
            await create_notification(
                admin_user['id'],
                f"Bill #{req.billNo} added to order {order['orderId']} by {user['firstName']}",
                "bill_added",
                order['id']
            )

    await create_audit_log(user['id'], "BILL_UPDATED", order_id, f"Bill {req.billNo} set on {order['orderId']}")
    await manager.broadcast({"type": "ORDER_UPDATED", "order": updated})
    return updated


@api_router.put("/orders/{order_id}/complete")
async def complete_order(order_id: str, user: dict = Depends(get_auth_user)):
    require_admin(user)
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    update = {
        "completed": True,
        "readinessStatus": "Completed",
        "completedAt": datetime.now(timezone.utc).isoformat(),
        "updatedAt": datetime.now(timezone.utc).isoformat(),
    }

    await db.orders.update_one({"id": order_id}, {"$set": update})
    updated = await db.orders.find_one({"id": order_id}, {"_id": 0})
    updated = await refresh_pdf_for_order(updated)
    await create_audit_log(user['id'], "ORDER_COMPLETED", order_id, f"Completed order {order['orderId']}")

    # Notify all users about order completion
    admin_users = await db.users.find({"role": "admin"}, {"_id": 0}).to_list(10)
    for admin_user in admin_users:
        await create_notification(admin_user['id'],
            f"Order {order['orderId']} has been completed!",
            "order_completed", order['id'])

    accountant_users = await db.users.find({"role": "accountant"}, {"_id": 0}).to_list(10)
    for accountant_user in accountant_users:
        await create_notification(accountant_user['id'],
            f"Order {order['orderId']} has been completed!",
            "order_completed", order['id'])

    staff_users = await db.users.find({"role": "staff"}, {"_id": 0}).to_list(100)
    for staff_user in staff_users:
        await create_notification(staff_user['id'],
            f"Order {order['orderId']} has been completed!",
            "order_completed", order['id'])

    await manager.broadcast({"type": "ORDER_UPDATED", "order": updated})
    return updated


@api_router.get("/orders/{order_id}/bill-pdf-url")
async def get_bill_pdf_url(order_id: str, user: dict = Depends(get_auth_user)):
    """Get the PDF bill URL for a completed order"""
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    if not order.get("billPdfUrl"):
        raise HTTPException(status_code=404, detail="PDF not available for this order")

    return {
        "orderId": order.get("orderId"),
        "billPdfUrl": order.get("billPdfUrl"),
        "fileName": f"{order.get('orderId')}_bill.pdf"
    }


@api_router.delete("/orders/{order_id}")
async def delete_order(order_id: str, user: dict = Depends(get_auth_user)):
    require_admin(user)
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    await db.orders.delete_one({"id": order_id})
    await create_audit_log(user['id'], "ORDER_DELETED", order_id, f"Deleted order {order['orderId']}")
    await manager.broadcast({"type": "ORDER_DELETED", "orderId": order_id})
    return {"message": "Order deleted"}


# ─── Order Status Routes ─────────────────────────────────────────────────────

@api_router.put("/orders/{order_id}/godown")
async def update_godown(order_id: str, req: GodownUpdateRequest, user: dict = Depends(get_auth_user)):
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    godown_dist = order.get('godownDistribution', [])
    found = False
    for g in godown_dist:
        if g['godown'] == req.godown:
            g['readyParcels'] = req.readyParcels
            found = True
            break
    if not found:
        godown_dist.append({"godown": req.godown, "readyParcels": req.readyParcels})
    total_ready = sum(g['readyParcels'] for g in godown_dist)
    if total_ready >= order['totalParcels']:
        readiness = "Ready"
    elif total_ready > 0:
        readiness = "Partial Ready"
    else:
        readiness = "Pending"
    await db.orders.update_one({"id": order_id}, {"$set": {
        "godownDistribution": godown_dist,
        "readinessStatus": readiness,
        "updatedAt": datetime.now(timezone.utc).isoformat()
    }})
    updated = await db.orders.find_one({"id": order_id}, {"_id": 0})
    updated = await refresh_pdf_for_order(updated)
    await create_audit_log(user['id'], "GODOWN_UPDATED", order_id,
        f"Godown {req.godown}: {req.readyParcels} parcels for {order['orderId']}")
    if readiness == "Ready":
        admins = await db.users.find({"role": "admin"}, {"_id": 0}).to_list(10)
        for admin_user in admins:
            await create_notification(admin_user['id'],
                f"Order {order['orderId']} is fully ready for dispatch!",
                "order_ready", order['id'])
    await manager.broadcast({"type": "ORDER_UPDATED", "order": updated})
    return updated


@api_router.put("/orders/{order_id}/dispatch")
async def toggle_dispatch(order_id: str, request: Request, user: dict = Depends(get_auth_user)):
    try:
        body = await request.json()
        dispatch_note = body.get("dispatchNote", "")
    except Exception:
        dispatch_note = ""

    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    new_val = not order.get('dispatched', False)

    update_data = {
        "dispatched": new_val,
        "updatedAt": datetime.now(timezone.utc).isoformat()
    }
    if new_val:
        update_data["dispatchedAt"] = datetime.now(timezone.utc).isoformat()
        if dispatch_note:
            update_data["dispatchNote"] = dispatch_note

    await db.orders.update_one({"id": order_id}, {"$set": update_data})
    updated = await db.orders.find_one({"id": order_id}, {"_id": 0})
    updated = await refresh_pdf_for_order(updated)

    await create_audit_log(user['id'], "DISPATCH_UPDATED", order_id,
        f"Order {order['orderId']} {'dispatched' if new_val else 'un-dispatched'}")

    if new_val:
        # Notify admins and accountants of dispatch
        admin_users = await db.users.find({"role": "admin"}, {"_id": 0}).to_list(10)
        for admin_user in admin_users:
            await create_notification(admin_user['id'],
                f"Order {order['orderId']} has been dispatched!",
                "order_dispatched", order['id'])

        accountant_users = await db.users.find({"role": "accountant"}, {"_id": 0}).to_list(10)
        for accountant_user in accountant_users:
            await create_notification(accountant_user['id'],
                f"Order {order['orderId']} has been dispatched!",
                "order_dispatched", order['id'])

    await manager.broadcast({"type": "ORDER_UPDATED", "order": updated})
    return updated


@api_router.put("/orders/{order_id}/split")
async def split_order(order_id: str, req: Request, user: dict = Depends(get_auth_user)):
    """Split order into ready and remainder orders"""
    try:
        body = await req.json()
        remainder_godown = body.get("remainderGodown", "")
    except Exception:
        remainder_godown = ""

    if not remainder_godown or not remainder_godown.strip():
        raise HTTPException(status_code=400, detail="remainderGodown is required")

    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    items = order.get('items', [])
    remaining_items = []

    # Calculate remaining parcels for each item
    for item in items:
        fulfilled = sum(1 for w in item.get('fulfillment', []) if w is not None)
        remaining = item.get('quantity', 0) - fulfilled

        # Only add to remaining if there are unfulfilled parcels
        if remaining > 0:
            remaining_items.append({
                "productId": item['productId'],
                "alias": item['alias'],
                "category": item['category'],
                "size": item['size'],
                "printName": item['printName'],
                "quantity": remaining,
                "rate": item.get('rate'),
                "requireSerialNo": item.get('requireSerialNo', False)
            })

    if not remaining_items:
        raise HTTPException(status_code=400, detail="No unfulfilled parcels to split")

    # Create remainder order — use the destination godown's own numbering
    prefix, order_num = await get_next_order_id(remainder_godown)
    new_order_id_value = str(uuid.uuid4())
    new_order_display_id = f"{prefix}-{order_num:04d}"

    new_order = {
        "id": new_order_id_value,
        "orderId": new_order_display_id,
        "partyName": order['partyName'],
        "location": order.get('location', ''),
        "godown": remainder_godown,
        "message": f"Remaining parcels from {order['orderId']}",
        "totalParcels": sum(item['quantity'] for item in remaining_items),
        "items": remaining_items,
        "readinessStatus": "Pending",
        "dispatched": False,
        "dispatchedAt": None,
        "createdByName": user.get('name', 'System'),
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "updatedAt": datetime.now(timezone.utc).isoformat(),
        "godownDistribution": [],
        "parentOrderId": order_id
    }

    # Insert the new order
    result = await db.orders.insert_one(new_order)

    # Fetch the newly created order to get it with proper formatting
    inserted_order = await db.orders.find_one({"id": new_order_id_value}, {"_id": 0})

    # Update original order: keep only fulfilled items, update totalParcels
    original_order_items = order.get('items', [])

    # Create updated items list with only fulfilled parcels
    updated_items_for_original = []
    total_fulfilled_in_original = 0

    for item in original_order_items:
        fulfilled_weights = [w for w in item.get('fulfillment', []) if w is not None]
        fulfilled = len(fulfilled_weights)

        if fulfilled > 0:
            # Keep only fulfilled parcels in original order
            updated_item = {
                "productId": item['productId'],
                "alias": item['alias'],
                "category": item['category'],
                "size": item['size'],
                "printName": item['printName'],
                "quantity": fulfilled,
                "rate": item.get('rate'),
                "requireSerialNo": item.get('requireSerialNo', False),
                "fulfillment": fulfilled_weights,
            }
            if item.get('serialNumbers'):
                # Keep only fulfilled serial numbers
                updated_item['serialNumbers'] = [
                    item['serialNumbers'][i]
                    for i in range(len(item.get('fulfillment', [])))
                    if item['fulfillment'][i] is not None
                ]
            updated_items_for_original.append(updated_item)
            total_fulfilled_in_original += fulfilled

    new_original_status = "Ready" if total_fulfilled_in_original > 0 else "Pending"
    original_order_update = {
        "items": updated_items_for_original,
        "totalParcels": total_fulfilled_in_original,
        "readinessStatus": new_original_status,
        "updatedAt": datetime.now(timezone.utc).isoformat()
    }

    # Recalculate godownDistribution for original order
    godown_dist = {}
    for item in updated_items_for_original:
        item_godown = order.get('godown', '')
        if item_godown not in godown_dist:
            godown_dist[item_godown] = 0
        godown_dist[item_godown] += item['quantity']

    original_order_update["godownDistribution"] = [
        {"godown": godown, "readyParcels": parcels}
        for godown, parcels in godown_dist.items()
    ]

    await db.orders.update_one({"id": order_id}, {"$set": original_order_update})
    updated_original = await db.orders.find_one({"id": order_id}, {"_id": 0})
    updated_original = await refresh_pdf_for_order(updated_original)

    # Notify admins and accountants about the split order
    admins = await db.users.find({"role": "admin"}, {"_id": 0}).to_list(10)
    for admin_user in admins:
        await create_notification(admin_user['id'],
            f"Order {order['orderId']} split into 2 orders. Remainder {new_order_display_id} created in {remainder_godown}",
            "order_created", new_order_id_value)

    accountants = await db.users.find({"role": "accountant"}, {"_id": 0}).to_list(10)
    for accountant_user in accountants:
        await create_notification(accountant_user['id'],
            f"Order {order['orderId']} split into 2 orders. Remainder {new_order_display_id} created in {remainder_godown}",
            "order_created", new_order_id_value)

    await create_audit_log(user['id'], "SPLIT_ORDER", order_id,
        f"Order {order['orderId']} split. Remainder: {new_order_display_id}")

    # Broadcast both updated orders
    await manager.broadcast({"type": "ORDER_UPDATED", "order": updated_original})
    await manager.broadcast({"type": "ORDER_CREATED", "order": inserted_order})

    return {"success": True, "originalOrder": updated_original, "remainderOrder": inserted_order}


@api_router.put("/orders/{order_id}/fulfill")
async def fulfill_parcel(order_id: str, req: ParcelFulfillmentRequest, user: dict = Depends(get_auth_user)):
    logger.info(f"Fulfill request: order_id={order_id}, productId={req.productId}, parcelIndex={req.parcelIndex}, weight={req.weight}")

    # Try to find order by id field
    order = await db.orders.find_one({"id": order_id})
    if not order:
        # Fallback: try orderId
        order = await db.orders.find_one({"orderId": order_id})

    if not order:
        # Final fallback: try _id as ObjectId
        try:
            order = await db.orders.find_one({"_id": ObjectId(order_id)})
        except:
            logger.warning(f"Could not parse {order_id} as ObjectId")

    if not order:
        logger.error(f"Order not found with any field matching: {order_id}")
        raise HTTPException(status_code=404, detail=f"Order not found")

    items = order.get('items', [])
    updated_items = []
    item_found = False

    for item in items:
        if item['productId'] == req.productId:
            item_found = True
            # Initialize fulfillment array if not present or None
            if not item.get('fulfillment'):
                item['fulfillment'] = [None] * item['quantity']

            # Ensure array has enough slots
            while len(item['fulfillment']) < item['quantity']:
                item['fulfillment'].append(None)

            # Update parcel weight (always save if in request)
            if req.parcelIndex < len(item['fulfillment']):
                item['fulfillment'][req.parcelIndex] = req.weight

            # Update parcel serial number (always save if in request, convert empty string to None)
            if not item.get('serialNumbers'):
                item['serialNumbers'] = [None] * item['quantity']

            while len(item['serialNumbers']) < item['quantity']:
                item['serialNumbers'].append(None)

            if req.parcelIndex < len(item['serialNumbers']):
                item['serialNumbers'][req.parcelIndex] = req.serialNo if req.serialNo and req.serialNo.strip() else None

        updated_items.append(item)

    if not item_found:
        logger.error(f"Product {req.productId} not found in order {order_id}. Available products: {[item.get('productId') for item in items]}")
        raise HTTPException(status_code=404, detail=f"Product {req.productId} not found in order")

    # Calculate status based on fulfillment
    total_parcels = order.get('totalParcels', 0)
    fulfilled_count = 0
    for item in updated_items:
        fulfilled = sum(1 for w in item.get('fulfillment', []) if w is not None)
        fulfilled_count += fulfilled

    new_status = "Pending"
    if fulfilled_count == total_parcels:
        new_status = "Ready"
    elif fulfilled_count > 0:
        new_status = "Partial Ready"

    update_dict = {
        "items": updated_items,
        "readinessStatus": new_status,
        "updatedAt": datetime.now(timezone.utc).isoformat()
    }

    await db.orders.update_one(
        {"_id": order["_id"]},
        {"$set": update_dict}
    )

    updated = await db.orders.find_one({"_id": order["_id"]}, {"_id": 0})
    updated = await refresh_pdf_for_order(updated)
    await create_audit_log(user['id'], "PARCEL_FULFILLED", order_id,
        f"Parcel {req.parcelIndex + 1} for {req.productId} fulfilled with {req.weight}kg")
    await manager.broadcast({"type": "ORDER_UPDATED", "order": updated})
    return updated


# ─── Product Routes ───────────────────────────────────────────────────────────

@api_router.get("/products/categories")
async def get_categories(user: dict = Depends(get_auth_user)):
    categories = await db.products.distinct("category")
    # Enrich with metadata
    metas = await db.category_meta.find({}, {"_id": 0}).to_list(500)
    meta_map = {m["name"]: m for m in metas}
    result = []
    for cat in sorted(categories):
        meta = meta_map.get(cat, {})
        result.append({
            "name": cat,
            "requireSerialNo": meta.get("requireSerialNo", False),
        })
    return result


class UpdateCategoryMetaRequest(BaseModel):
    requireSerialNo: Optional[bool] = None
    newName: Optional[str] = None


@api_router.put("/products/categories/{category_name}")
async def update_category_meta(category_name: str, req: UpdateCategoryMetaRequest, user: dict = Depends(get_auth_user)):
    require_admin(user)
    update_meta: dict = {}
    if req.requireSerialNo is not None:
        update_meta["requireSerialNo"] = req.requireSerialNo
    if req.newName and req.newName.strip() and req.newName.strip() != category_name:
        new_name = req.newName.strip()
        # Rename all products in this category
        await db.products.update_many({"category": category_name}, {"$set": {"category": new_name}})
        # Update old meta doc name
        await db.category_meta.delete_one({"name": category_name})
        category_name = new_name
    await db.category_meta.update_one(
        {"name": category_name},
        {"$set": {**update_meta, "name": category_name}},
        upsert=True
    )
    return {"name": category_name, **update_meta}


@api_router.get("/products")
async def get_products(
    user: dict = Depends(get_auth_user),
    category: Optional[str] = None,
    search: Optional[str] = None
):
    query = {}
    if category:
        query["category"] = category
    if search:
        query["$or"] = [
            {"category": {"$regex": search, "$options": "i"}},
            {"size": {"$regex": search, "$options": "i"}},
            {"printName": {"$regex": search, "$options": "i"}},
            {"alias": {"$regex": search, "$options": "i"}}
        ]
    products = await db.products.find(query, {"_id": 0}).sort("category", 1).to_list(1000)
    return products


@api_router.post("/products")
async def create_product(req: CreateProductRequest, user: dict = Depends(get_auth_user)):
    require_admin(user)
    existing = await db.products.find_one({"alias": req.alias})
    if existing:
        raise HTTPException(status_code=400, detail="Product alias already exists")
    product = {
        "id": str(uuid.uuid4()),
        "category": req.category,
        "size": req.size,
        "printName": req.printName,
        "alias": req.alias,
        "createdAt": datetime.now(timezone.utc).isoformat(),
        "updatedAt": datetime.now(timezone.utc).isoformat()
    }
    await db.products.insert_one({**product})
    await create_audit_log(user['id'], "PRODUCT_CREATED", None, f"Created product {req.alias}")
    return product


@api_router.post("/products/import")
async def import_products(
    file: UploadFile = File(...),
    user: dict = Depends(get_auth_user),
):
    """Bulk import products from an Excel file.

    The file must match the format of ListofItems.xlsx used for seeding:
    Sheet1, data starts at row 4, columns A..D are
    category | size | printName | alias. Rows whose alias already exists
    in the database are left untouched (counted as duplicates); only new
    aliases are inserted. Returns a summary of rows processed.
    """
    require_admin(user)

    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be an Excel spreadsheet (.xlsx / .xls)")

    try:
        import openpyxl
        from io import BytesIO
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel import library not available on the server")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    try:
        wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read the spreadsheet: {e}")

    # Prefer Sheet1 (matches ListofItems.xlsx), but fall back to the first
    # sheet so users don't have to rename tabs before uploading.
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    # Detect the data start row. ListofItems.xlsx uses rows 1-3 as header
    # metadata and row 4 onwards for data. If the first row already looks
    # like data (column A is non-empty and not a label), start at row 1.
    start_row = 4
    first_val = ws.cell(1, 1).value
    if first_val and str(first_val).strip().lower() not in {
        "list of items", "category", "item category", "items", "s.no"
    }:
        start_row = 1

    inserted = 0
    duplicates = 0
    skipped = 0
    errors: List[str] = []
    now_iso = datetime.now(timezone.utc).isoformat()

    # Snapshot the existing aliases up-front so we don't hit the DB once per
    # row. The catalog is small enough (a few hundred entries) that holding
    # the set in memory is cheap and saves O(N) round-trips on big imports.
    existing_alias_set = {
        doc["alias"]
        for doc in await db.products.find({}, {"_id": 0, "alias": 1}).to_list(length=None)
        if doc.get("alias")
    }

    for row_idx in range(start_row, ws.max_row + 1):
        category = ws.cell(row_idx, 1).value
        size = ws.cell(row_idx, 2).value
        print_name = ws.cell(row_idx, 3).value
        alias = ws.cell(row_idx, 4).value

        # Blank/header-ish rows — silently skip.
        if not category or not size or not alias:
            skipped += 1
            continue

        category_s = str(category).strip()
        size_s = str(size).strip()
        print_name_s = str(print_name).strip() if print_name else ""
        alias_s = str(alias).strip()

        if not category_s or not size_s or not alias_s:
            skipped += 1
            continue

        # If the alias already lives in the catalog, leave it alone — the
        # admin asked us not to overwrite existing products. Track it as a
        # duplicate so the import summary still tells them what we saw.
        if alias_s in existing_alias_set:
            duplicates += 1
            continue

        try:
            await db.products.insert_one({
                "id": str(uuid.uuid4()),
                "category": category_s,
                "size": size_s,
                "printName": print_name_s,
                "alias": alias_s,
                "createdAt": now_iso,
                "updatedAt": now_iso,
            })
            existing_alias_set.add(alias_s)
            inserted += 1
        except Exception as e:
            errors.append(f"Row {row_idx} ({alias_s}): {e}")

    await create_audit_log(
        user['id'],
        "PRODUCTS_IMPORTED",
        None,
        f"Imported products from {file.filename}: {inserted} new, {duplicates} duplicates, {skipped} skipped",
    )
    return {
        "inserted": inserted,
        "duplicates": duplicates,
        "skipped": skipped,
        "errors": errors[:20],
    }


@api_router.get("/products/{product_id}")
async def get_product(product_id: str, user: dict = Depends(get_auth_user)):
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product


@api_router.put("/products/{product_id}")
async def update_product(product_id: str, req: UpdateProductRequest, user: dict = Depends(get_auth_user)):
    require_admin(user)
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    update = {"updatedAt": datetime.now(timezone.utc).isoformat()}
    if req.category is not None:
        update["category"] = req.category
    if req.size is not None:
        update["size"] = req.size
    if req.printName is not None:
        update["printName"] = req.printName
    if req.alias is not None:
        update["alias"] = req.alias
    await db.products.update_one({"id": product_id}, {"$set": update})
    updated = await db.products.find_one({"id": product_id}, {"_id": 0})
    await create_audit_log(user['id'], "PRODUCT_UPDATED", None, f"Updated product {product['alias']}")
    return updated


@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, user: dict = Depends(get_auth_user)):
    require_admin(user)
    product = await db.products.find_one({"id": product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    await db.products.delete_one({"id": product_id})
    await create_audit_log(user['id'], "PRODUCT_DELETED", None, f"Deleted product {product['alias']}")
    return {"message": "Product deleted"}


# ─── Gowdown Routes ──────────────────────────────────────────────────────────

@api_router.get("/gowdowns")
async def get_gowdowns(user: dict = Depends(get_auth_user)):
    gowdowns = await db.gowdowns.find({}, {"_id": 0}).to_list(100)
    return gowdowns


@api_router.put("/gowdowns/{gowdown_id}/prefix")
async def update_gowdown_prefix(
    gowdown_id: str,
    req: GodownPrefixRequest,
    user: dict = Depends(get_auth_user),
):
    require_admin(user)
    # Normalise the prefix: uppercase, letters/digits only, 1–5 chars.
    prefix = re.sub(r"[^A-Za-z0-9]", "", req.prefix or "").upper()
    if not prefix or len(prefix) > 5:
        raise HTTPException(status_code=400, detail="Prefix must be 1–5 alphanumeric characters")

    godown = await db.gowdowns.find_one({"id": gowdown_id})
    if not godown:
        raise HTTPException(status_code=404, detail="Gowdown not found")

    # Two godowns sharing the same prefix would collide on the counter.
    clash = await db.gowdowns.find_one({"prefix": prefix, "id": {"$ne": gowdown_id}})
    if clash:
        raise HTTPException(status_code=409, detail=f"Prefix '{prefix}' is already used by {clash['name']}")

    await db.gowdowns.update_one(
        {"id": gowdown_id},
        {"$set": {"prefix": prefix, "updatedAt": datetime.now(timezone.utc).isoformat()}},
    )
    await create_audit_log(user['id'], "GODOWN_PREFIX_UPDATED", gowdown_id, f"Prefix set to {prefix}")
    return await db.gowdowns.find_one({"id": gowdown_id}, {"_id": 0})


# ─── Notification Routes ─────────────────────────────────────────────────────

@api_router.get("/notifications")
async def get_notifications(user: dict = Depends(get_auth_user)):
    notifs = await db.notifications.find(
        {"userId": user['id']}, {"_id": 0}
    ).sort("createdAt", -1).to_list(100)
    return notifs


@api_router.get("/notifications/unread-count")
async def get_unread_count(user: dict = Depends(get_auth_user)):
    count = await db.notifications.count_documents({"userId": user['id'], "read": False})
    return {"count": count}


@api_router.put("/notifications/{notification_id}/read")
async def mark_notification_read(notification_id: str, user: dict = Depends(get_auth_user)):
    await db.notifications.update_one(
        {"id": notification_id, "userId": user['id']},
        {"$set": {"read": True}}
    )
    return {"message": "Marked as read"}


@api_router.put("/notifications/read-all")
async def mark_all_read(user: dict = Depends(get_auth_user)):
    await db.notifications.update_many(
        {"userId": user['id'], "read": False},
        {"$set": {"read": True}}
    )
    return {"message": "All marked as read"}


# ─── Dashboard & Audit ────────────────────────────────────────────────────────

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(user: dict = Depends(get_auth_user)):
    user_role = user.get('role', 'staff')
    non_completed = {"completed": {"$ne": True}}

    if user_role == 'admin':
        # Admin: active = all non-completed
        total_active = await db.orders.count_documents(non_completed)
        pending = await db.orders.count_documents({**non_completed, "readinessStatus": "Pending"})
        partial = await db.orders.count_documents({**non_completed, "readinessStatus": "Partial Ready"})
        ready = await db.orders.count_documents({**non_completed, "readinessStatus": "Ready"})
        dispatched = await db.orders.count_documents({**non_completed, "dispatched": True, "readinessStatus": {"$ne": "Bill Generated"}})
        bill_generated = await db.orders.count_documents({**non_completed, "readinessStatus": "Bill Generated"})
        completed = await db.orders.count_documents({"completed": True})
        return {
            "totalActive": total_active,
            "pending": pending,
            "partialReady": partial,
            "ready": ready,
            "dispatched": dispatched,
            "billGenerated": bill_generated,
            "completed": completed,
        }
    elif user_role == 'accountant':
        # Accountant: needs-bill = orders that are Ready or dispatched and have no bill number yet
        needs_bill = await db.orders.count_documents({
            **non_completed,
            "$and": [
                {"$or": [{"readinessStatus": "Ready"}, {"dispatched": True}]},
                {"$or": [{"billNo": None}, {"billNo": {"$exists": False}}, {"billNo": ""}]},
            ],
        })
        bill_generated = await db.orders.count_documents({**non_completed, "readinessStatus": "Bill Generated"})
        return {
            "totalActive": needs_bill,
            "needsBill": needs_bill,
            "billGenerated": bill_generated,
        }
    else:
        # Staff: active = pending + partial + ready (non-dispatched)
        base = {"dispatched": False, **non_completed}
        total = await db.orders.count_documents(base)
        pending = await db.orders.count_documents({**base, "readinessStatus": "Pending"})
        partial = await db.orders.count_documents({**base, "readinessStatus": "Partial Ready"})
        ready = await db.orders.count_documents({**base, "readinessStatus": "Ready"})
        return {
            "totalActive": total,
            "pending": pending,
            "partialReady": partial,
            "ready": ready,
        }


@api_router.get("/audit-logs")
async def get_audit_logs(user: dict = Depends(get_auth_user), limit: int = 50):
    require_admin(user)
    logs = await db.audit_logs.find({}, {"_id": 0}).sort("timestamp", -1).to_list(limit)
    return logs


# ─── WebSocket ────────────────────────────────────────────────────────────────

@api_router.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket, token: str = Query("")):
    if not token:
        logger.warning("WebSocket connection rejected: no token provided")
        await websocket.close(code=4001)
        return
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload['userId']}, {"_id": 0})
        if not user:
            logger.warning(f"WebSocket connection rejected: user not found {payload.get('userId')}")
            await websocket.close(code=4001)
            return
    except jwt.ExpiredSignatureError:
        logger.warning("WebSocket connection rejected: token expired")
        await websocket.close(code=4001)
        return
    except jwt.InvalidTokenError as e:
        logger.warning(f"WebSocket connection rejected: invalid token - {str(e)}")
        await websocket.close(code=4001)
        return
    except Exception as e:
        logger.error(f"WebSocket connection error: {str(e)}")
        await websocket.close(code=4001)
        return

    logger.info(f"WebSocket connection accepted for user {user['id']}")
    await manager.connect(websocket, user['id'])
    try:
        while True:
            data = await websocket.receive_text()
            if data == "ping":
                await websocket.send_text("pong")
    except WebSocketDisconnect:
        manager.disconnect(user['id'], websocket)
    except Exception as e:
        logger.error(f"WebSocket error for user {user['id']}: {str(e)}")
        manager.disconnect(user['id'], websocket)


# ─── Startup ──────────────────────────────────────────────────────────────────

async def seed_products_from_xlsx():
    """Load and seed products from ListofItems.xlsx if not already seeded."""
    try:
        import openpyxl
        xlsx_path = Path(__file__).parent.parent / 'ListofItems.xlsx'

        if not xlsx_path.exists():
            logger.warning(f"ListofItems.xlsx not found at {xlsx_path}")
            return

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb['Sheet1']

        products = []
        for row in range(4, ws.max_row + 1):
            category = ws.cell(row, 1).value
            size = ws.cell(row, 2).value
            print_name = ws.cell(row, 3).value
            alias = ws.cell(row, 4).value

            if not category or not size or not alias:
                continue

            products.append({
                "id": str(uuid.uuid4()),
                "category": str(category).strip(),
                "size": str(size).strip(),
                "printName": str(print_name).strip() if print_name else "",
                "alias": str(alias).strip(),
                "createdAt": datetime.now(timezone.utc).isoformat(),
                "updatedAt": datetime.now(timezone.utc).isoformat(),
            })

        if products:
            # Use upsert to avoid duplicates
            for product in products:
                await db.products.update_one(
                    {"alias": product["alias"]},
                    {"$set": product},
                    upsert=True
                )
            logger.info(f"Seeded {len(products)} products from xlsx")
    except Exception as e:
        logger.warning(f"Failed to seed products: {e}")


@app.on_event("startup")
async def startup():
    await db.orders.create_index("orderId")
    await db.orders.create_index("createdAt")
    await db.orders.create_index("dispatched")
    await db.users.create_index("phone", unique=True)
    await db.notifications.create_index([("userId", 1), ("createdAt", -1)])
    await db.products.create_index("category")
    await db.products.create_index("alias", unique=True)

    # Product seeding disabled - use manual import for original data
    product_count = await db.products.count_documents({})
    logger.info(f"Products in database: {product_count}")

    # Seed gowdowns with default per-gowdown order-id prefixes. Admins can
    # override these later via PUT /gowdowns/{id}/prefix.
    default_gowdowns = [
        {"name": "Sundha", "prefix": "SU"},
        {"name": "Lal-Shivnagar", "prefix": "LS"},
    ]
    for gd in default_gowdowns:
        existing = await db.gowdowns.find_one({"name": gd["name"]})
        now_iso = datetime.now(timezone.utc).isoformat()
        if not existing:
            await db.gowdowns.insert_one({
                "id": str(uuid.uuid4()),
                "name": gd["name"],
                "prefix": gd["prefix"],
                "createdAt": now_iso,
                "updatedAt": now_iso,
            })
        elif not existing.get("prefix"):
            await db.gowdowns.update_one(
                {"id": existing["id"]},
                {"$set": {"prefix": gd["prefix"], "updatedAt": now_iso}},
            )
    logger.info("Gowdowns seeded successfully")

    counter = await db.counters.find_one({"name": "orderId"})
    if not counter:
        await db.counters.insert_one({"name": "orderId", "value": 0})
    logger.info("Database seeded successfully")


app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(api_router)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
