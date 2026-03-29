"""
Seed products from ListofItems.xlsx into MongoDB.
Reads the xlsx, creates product entries matching the backend schema.
"""
import asyncio
import uuid
import os
from datetime import datetime, timezone
import openpyxl
from motor.motor_asyncio import AsyncIOMotorClient
from dotenv import load_dotenv

load_dotenv()

XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', 'ListofItems.xlsx')
MONGO_URL = os.environ.get('MONGODB_URL')
DB_NAME = os.environ.get('DB_NAME', 'kishore_hosiery')


async def seed():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb['Sheet1']

    # Row 2 has headers: Category, Size, Print Name, Alias
    # Data starts at row 4
    products = []
    for row in range(4, ws.max_row + 1):
        category = ws.cell(row, 1).value
        size = ws.cell(row, 2).value
        print_name = ws.cell(row, 3).value
        alias = ws.cell(row, 4).value

        # Skip empty rows
        if not category or not size or not alias:
            continue

        products.append({
            "id": str(uuid.uuid4()),
            "category": str(category).strip(),
            "size": str(size).strip(),
            "printName": str(print_name).strip() if print_name else "",
            "alias": str(alias).strip(),
            "createdAt": datetime.now(timezone.utc).isoformat(),
            "updatedAt": datetime.now(timezone.utc).isoformat(),
        })

    print(f"Found {len(products)} products in xlsx")

    # Connect to MongoDB
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]

    # Clear existing products
    deleted = await db.products.delete_many({})
    print(f"Cleared {deleted.deleted_count} existing products")

    # Insert all
    if products:
        await db.products.insert_many(products)
        print(f"Inserted {len(products)} products")

    # Print unique categories
    categories = sorted(set(p['category'] for p in products))
    print(f"\n{len(categories)} categories:")
    for c in categories:
        count = sum(1 for p in products if p['category'] == c)
        print(f"  {c} ({count} variants)")

    client.close()


if __name__ == '__main__':
    asyncio.run(seed())
